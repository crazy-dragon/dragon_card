import math
import os
import io
import json
import re
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file
from models import db, User, Template, Deck, DeckTemplate, DeckItem, Progress, StudyRound, LearningEvent
from config import Config

MAX_ACTIONS = 5
DECK_KINDS = {'language', 'knowledge', 'logic', 'skill', 'other'}


def create_app():
    app = Flask(__name__, static_folder='static')
    app.config.from_object(Config)
    db.init_app(app)

    with app.app_context():
        db.create_all()
        _ensure_default_user()
        _migrate_schema()

    return app


def _ensure_default_user():
    if not User.query.filter_by(username='default').first():
        db.session.add(User(username='default'))
        db.session.commit()


def _migrate_schema():
    from sqlalchemy import inspect
    inspector = inspect(db.engine)

    tables = inspector.get_table_names()

    if 't_deck' not in tables:
        db.session.execute(db.text('CREATE TABLE t_deck (id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL REFERENCES t_user(id), name VARCHAR(100) NOT NULL, kind VARCHAR(20) DEFAULT \'other\', active_template_id INTEGER REFERENCES t_template(id), created_at DATETIME DEFAULT CURRENT_TIMESTAMP, updated_at DATETIME DEFAULT CURRENT_TIMESTAMP)'))
        db.session.commit()

    if 't_deck_template' not in tables:
        db.session.execute(db.text('CREATE TABLE t_deck_template (deck_id INTEGER NOT NULL REFERENCES t_deck(id), template_id INTEGER NOT NULL REFERENCES t_template(id), sort_order INTEGER DEFAULT 0, PRIMARY KEY (deck_id, template_id))'))
        db.session.commit()

    if 't_deck_item' not in tables:
        db.session.execute(db.text('CREATE TABLE t_deck_item (id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, deck_id INTEGER NOT NULL REFERENCES t_deck(id), item_order INTEGER NOT NULL, data JSON NOT NULL, debug BOOLEAN DEFAULT 0, created_at DATETIME DEFAULT CURRENT_TIMESTAMP, updated_at DATETIME DEFAULT CURRENT_TIMESTAMP)'))
        db.session.commit()

    deck_cols = [c['name'] for c in inspector.get_columns('t_deck')]
    if 'kind' not in deck_cols:
        db.session.execute(db.text("ALTER TABLE t_deck ADD COLUMN kind VARCHAR(20) DEFAULT 'other'"))
        db.session.commit()
    if 'description' in deck_cols:
        db.session.execute(db.text('ALTER TABLE t_deck DROP COLUMN description'))
        db.session.commit()

    tmpl_cols = [c['name'] for c in inspector.get_columns('t_template')]
    if 'tracked_actions' not in tmpl_cols:
        db.session.execute(db.text('ALTER TABLE t_template ADD COLUMN tracked_actions TEXT'))
        db.session.commit()
    if 'lang' not in tmpl_cols:
        db.session.execute(db.text("ALTER TABLE t_template ADD COLUMN lang VARCHAR(10) DEFAULT 'en'"))
        db.session.commit()

    user_cols = [c['name'] for c in inspector.get_columns('t_user')]
    if 'display_name' not in user_cols:
        db.session.execute(db.text('ALTER TABLE t_user ADD COLUMN display_name VARCHAR(100)'))
        db.session.commit()


app = create_app()


# ==================== User ====================

@app.route('/v1/users/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    if not username:
        return jsonify({'error': 'username required'}), 400
    user = User.query.filter_by(username=username).first()
    if not user:
        user = User(username=username)
        db.session.add(user)
        db.session.commit()
    return jsonify({'success': True, 'user': user.to_dict()})


@app.route('/v1/users')
def list_users():
    users = User.query.all()
    return jsonify({'success': True, 'users': [u.to_dict() for u in users]})


@app.route('/v1/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    u = db.session.get(User, user_id)
    if not u:
        return jsonify({'error': 'User not found'}), 404
    data = request.get_json()
    if 'display_name' in data:
        u.display_name = data['display_name']
    if 'username' in data:
        existing = User.query.filter(User.username == data['username'], User.id != user_id).first()
        if existing:
            return jsonify({'error': 'Username already taken'}), 400
        u.username = data['username']
    db.session.commit()
    return jsonify({'success': True, 'user': u.to_dict()})


# ==================== Template ====================

def _count_template_actions(js_text):
    """Count distinct data-action values in card_html + card_js."""
    actions = set()
    for m in re.finditer(r'data-action=["\']([^"\']+)["\']', js_text):
        actions.add(m.group(1))
    return len(actions)


@app.route('/v1/templates', methods=['GET'])
def list_templates():
    user_id = request.args.get('user_id', type=int)
    q = Template.query
    if user_id:
        q = q.filter((Template.user_id == user_id) | (Template.user_id.is_(None)))
    templates = q.all()
    return jsonify({'success': True, 'templates': [t.to_dict() for t in templates]})


@app.route('/v1/templates', methods=['POST'])
def create_template():
    data = request.get_json()
    t = Template(
        user_id=data.get('user_id'),
        name=data.get('name', ''),
        description=data.get('description', ''),
        lang=data.get('lang', 'en') or 'en',
        card_html=data.get('card_html', ''),
        card_css=data.get('card_css', ''),
        card_js=data.get('card_js', ''),
    )
    db.session.add(t)
    db.session.commit()
    return jsonify({'success': True, 'template': t.to_dict()}), 201


@app.route('/v1/templates/<int:template_id>', methods=['GET'])
def get_template(template_id):
    t = db.session.get(Template, template_id)
    if not t:
        return jsonify({'error': 'Template not found'}), 404
    return jsonify({'success': True, 'template': t.to_dict_full()})


@app.route('/v1/templates/<int:template_id>', methods=['PUT'])
def update_template(template_id):
    t = db.session.get(Template, template_id)
    if not t:
        return jsonify({'error': 'Template not found'}), 404
    data = request.get_json()
    for field in ('name', 'description', 'lang', 'card_html', 'card_css', 'card_js', 'sample_data', 'tracked_actions'):
        if field in data:
            setattr(t, field, data[field])
    db.session.commit()
    return jsonify({'success': True, 'template': t.to_dict()})


@app.route('/v1/templates/<int:template_id>', methods=['DELETE'])
def delete_template(template_id):
    t = db.session.get(Template, template_id)
    if not t:
        return jsonify({'error': 'Template not found'}), 404

    backup_dir = os.path.join(os.path.dirname(__file__), 'backups', 'templates')
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, f'{t.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json')
    with open(backup_path, 'w', encoding='utf-8') as f:
        json.dump(t.to_dict_full(), f, ensure_ascii=False, indent=2)

    Deck.query.filter_by(active_template_id=template_id).update({'active_template_id': None})
    DeckTemplate.query.filter_by(template_id=template_id).delete()
    db.session.delete(t)
    db.session.commit()
    return jsonify({'success': True, 'backup': backup_path})


def _parse_template(text):
    """Parse a JSON-format template file."""
    data = json.loads(text.strip())
    sample_data = None
    if 'sampleData' in data:
        sample_data = json.dumps(data['sampleData'], ensure_ascii=False)
    tracked = data.get('trackedActions', [])
    if isinstance(tracked, list):
        tracked_actions = json.dumps(tracked, ensure_ascii=False)
    else:
        tracked_actions = ''
    return {
        'name': data.get('name', ''),
        'description': data.get('description', ''),
        'lang': data.get('lang', 'en') or 'en',
        'cardHtml': data.get('cardHtml', ''),
        'cardCss': data.get('cardCss', ''),
        'cardJs': data.get('cardJs', ''),
        'sampleData': sample_data or '',
        'trackedActions': tracked_actions,
    }


@app.route('/v1/templates/import', methods=['POST'])
def import_template():
    data = request.get_json()
    text = data.get('content', '')
    if not text:
        return jsonify({'error': 'No template content'}), 400

    parsed = _parse_template(text)
    if not parsed['name']:
        return jsonify({'error': 'Invalid template format: name not found'}), 400

    # Validate action count
    action_count = _count_template_actions(parsed['cardHtml'] + parsed['cardJs'])
    if action_count > MAX_ACTIONS:
        return jsonify({
            'error': f'Template has {action_count} actions, maximum is {MAX_ACTIONS}'
        }), 400

    sample_data_raw = parsed.get('sampleData', '').strip()
    if sample_data_raw:
        try:
            json.loads(sample_data_raw)
        except json.JSONDecodeError:
            return jsonify({'error': 'Invalid JSON in ==sampleData== section'}), 400

    t = Template(
        user_id=data.get('user_id'),
        name=parsed['name'],
        description=parsed['description'],
        lang=parsed.get('lang', 'en') or 'en',
        card_html=parsed['cardHtml'],
        card_css=parsed['cardCss'],
        card_js=parsed['cardJs'],
        sample_data=sample_data_raw or None,
        tracked_actions=parsed.get('trackedActions', ''),
    )
    db.session.add(t)
    db.session.commit()
    return jsonify({'success': True, 'template': t.to_dict()}), 201


@app.route('/v1/templates/<int:template_id>/export')
def export_template(template_id):
    t = db.session.get(Template, template_id)
    if not t:
        return jsonify({'error': 'Template not found'}), 404

    data = {
        'name': t.name,
        'description': t.description or '',
        'lang': t.lang or 'en',
        'cardHtml': t.card_html,
        'cardCss': t.card_css,
        'cardJs': t.card_js,
    }
    if t.sample_data:
        try:
            data['sampleData'] = json.loads(t.sample_data)
        except json.JSONDecodeError:
            pass

    return jsonify({
        'success': True,
        'name': t.name + '.json',
        'content': json.dumps(data, ensure_ascii=False, indent=2)
    })


# ==================== Deck (卡组) ====================

@app.route('/v1/decks', methods=['GET'])
def list_decks():
    user_id = request.args.get('user_id', type=int)
    q = Deck.query
    if user_id:
        q = q.filter_by(user_id=user_id)
    decks = q.order_by(Deck.created_at.desc()).all()

    # 各卡组今年有学习事件的天数（按卡组单独计算）
    year_days_map = _year_study_days_map(user_id) if user_id else {}

    result = []
    for d in decks:
        dd = d.to_dict()
        dd['year_study_days'] = year_days_map.get(d.id, 0)
        result.append(dd)
    return jsonify({'success': True, 'decks': result})


def _year_study_days_map(user_id):
    """各卡组今年有学习事件的天数（按 deck_id 分组，日期去重）。"""
    from datetime import date
    rows = db.session.query(
        LearningEvent.deck_id,
        db.func.count(db.func.distinct(db.func.date(LearningEvent.created_at)))
    ).filter(
        LearningEvent.user_id == user_id,
        db.func.date(LearningEvent.created_at) >= date(date.today().year, 1, 1),
    ).group_by(LearningEvent.deck_id).all()
    return {deck_id: days for deck_id, days in rows}


@app.route('/v1/decks', methods=['POST'])
def create_deck():
    data = request.get_json()
    kind = data.get('kind', 'other')
    if kind not in DECK_KINDS:
        kind = 'other'
    d = Deck(
        user_id=data['user_id'],
        name=data['name'],
        kind=kind,
    )
    db.session.add(d)
    db.session.commit()
    return jsonify({'success': True, 'deck': d.to_dict()}), 201


@app.route('/v1/decks/<int:deck_id>', methods=['GET'])
def get_deck(deck_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404
    return jsonify({'success': True, 'deck': d.to_dict()})


@app.route('/v1/decks/<int:deck_id>', methods=['PUT'])
def update_deck(deck_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404
    data = request.get_json()
    if 'name' in data:
        d.name = data['name']
    if 'kind' in data:
        kind = data['kind']
        if kind not in DECK_KINDS:
            return jsonify({'error': f'Invalid kind: {kind}'}), 400
        d.kind = kind
    db.session.commit()
    return jsonify({'success': True, 'deck': d.to_dict()})


@app.route('/v1/decks/<int:deck_id>', methods=['DELETE'])
def delete_deck(deck_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404
    DeckItem.query.filter_by(deck_id=deck_id).delete()
    Progress.query.filter_by(deck_id=deck_id).delete()
    StudyRound.query.filter_by(deck_id=deck_id).delete()
    LearningEvent.query.filter_by(deck_id=deck_id).delete()
    db.session.delete(d)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/v1/decks/<int:deck_id>/export')
def export_deck_data(deck_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404
    items = DeckItem.query.filter_by(deck_id=deck_id).order_by(DeckItem.item_order).all()
    data = [{
        'item_order': di.item_order,
        'data': di.data,
        'debug': di.debug,
    } for di in items]
    return jsonify({
        'success': True,
        'deck_name': d.name,
        'count': len(data),
        'data': data,
    })


@app.route('/v1/decks/<int:deck_id>/templates', methods=['POST'])
def upload_deck_template(deck_id):
    """Add (or replace) a template for a deck. Max 3."""
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404

    data = request.get_json()
    text = data.get('content', '')
    if not text:
        return jsonify({'error': 'No template content'}), 400

    parsed = _parse_template(text)
    if not parsed['name']:
        return jsonify({'error': 'Invalid template format: name not found'}), 400

    sample_data_raw = parsed.get('sampleData', '').strip()
    if sample_data_raw:
        try:
            json.loads(sample_data_raw)
        except json.JSONDecodeError:
            return jsonify({'error': 'Invalid JSON in sampleData'}), 400

    action_count = _count_template_actions(parsed['cardHtml'] + parsed['cardJs'])
    if action_count > MAX_ACTIONS:
        return jsonify({'error': f'Template has {action_count} actions, maximum is {MAX_ACTIONS}'}), 400

    replace_id = data.get('replace_template_id')

    if replace_id:
        t = db.session.get(Template, replace_id)
        if not t:
            return jsonify({'error': 'Template not found'}), 404
        if not DeckTemplate.query.filter_by(deck_id=deck_id, template_id=replace_id).first():
            return jsonify({'error': 'Template not linked to this deck'}), 400
        t.name = parsed['name']
        t.description = parsed['description']
        t.lang = parsed.get('lang', 'en') or 'en'
        t.card_html = parsed['cardHtml']
        t.card_css = parsed['cardCss']
        t.card_js = parsed['cardJs']
        t.sample_data = sample_data_raw or None
        t.tracked_actions = parsed.get('trackedActions', '')
    else:
        existing = DeckTemplate.query.filter_by(deck_id=deck_id).count()
        if existing >= 3:
            return jsonify({'error': 'Maximum 3 templates per deck'}), 400
        t = Template(
            user_id=d.user_id,
            name=parsed['name'],
            description=parsed['description'],
            lang=parsed.get('lang', 'en') or 'en',
            card_html=parsed['cardHtml'],
            card_css=parsed['cardCss'],
            card_js=parsed['cardJs'],
            sample_data=sample_data_raw or None,
            tracked_actions=parsed.get('trackedActions', ''),
        )
        db.session.add(t)
        db.session.flush()
        ct = DeckTemplate(deck_id=deck_id, template_id=t.id, sort_order=existing)
        db.session.add(ct)
        if not d.active_template_id:
            d.active_template_id = t.id

    d.active_template_id = t.id
    db.session.commit()
    return jsonify({'success': True, 'template': t.to_dict(), 'deck': d.to_dict()})


@app.route('/v1/decks/<int:deck_id>/preview')
def preview_deck(deck_id):
    """Preview a deck's template with its first data item. Optional ?template_id=N."""
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404
    tid = request.args.get('template_id', type=int) or d.active_template_id
    if not tid:
        return jsonify({'error': 'No template assigned to this deck'}), 400
    t = db.session.get(Template, tid)
    if not t:
        return jsonify({'error': 'Template not found'}), 404

    item = DeckItem.query.filter_by(deck_id=deck_id).order_by(DeckItem.item_order).first()
    sample_item = item.to_dict() if item else None

    if not sample_item and t.sample_data:
        try:
            import json
            sample_list = json.loads(t.sample_data)
            if sample_list:
                sample_item = {'id': 0, 'deck_id': deck_id, 'data': sample_list[0], 'is_unknown': 0, 'is_favorite': 0, 'item_order': 1, 'current_order': 1}
        except (json.JSONDecodeError, TypeError, IndexError):
            pass

    return jsonify({
        'success': True,
        'template': t.to_dict_full(),
        'sample_card': sample_item,
    })


@app.route('/v1/decks/<int:deck_id>/templates', methods=['GET'])
def list_deck_templates(deck_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404
    return jsonify({'success': True, 'templates': d.to_dict()['templates']})


@app.route('/v1/decks/<int:deck_id>/active-template', methods=['PUT'])
def set_active_deck_template(deck_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404
    data = request.get_json()
    tid = data.get('template_id')
    if not tid:
        return jsonify({'error': 'template_id required'}), 400
    if not DeckTemplate.query.filter_by(deck_id=deck_id, template_id=tid).first():
        return jsonify({'error': 'Template not linked to this deck'}), 400
    d.active_template_id = tid
    db.session.commit()
    return jsonify({'success': True, 'deck': d.to_dict()})


@app.route('/v1/decks/<int:deck_id>/templates/<int:template_id>', methods=['DELETE'])
def remove_deck_template(deck_id, template_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404
    dt = DeckTemplate.query.filter_by(deck_id=deck_id, template_id=template_id).first()
    if not dt:
        return jsonify({'error': 'Template not linked to this deck'}), 400
    db.session.delete(dt)
    if d.active_template_id == template_id:
        remaining = DeckTemplate.query.filter_by(deck_id=deck_id).order_by(DeckTemplate.sort_order).all()
        d.active_template_id = remaining[0].template_id if remaining else None
    db.session.commit()
    return jsonify({'success': True, 'deck': d.to_dict()})


# ==================== DeckItem (data entries) ====================

@app.route('/v1/decks/<int:deck_id>/items', methods=['GET'])
def list_deck_items(deck_id):
    items = DeckItem.query.filter_by(deck_id=deck_id).order_by(DeckItem.item_order).all()
    return jsonify({'success': True, 'items': [di.to_dict() for di in items]})


@app.route('/v1/decks/<int:deck_id>/import', methods=['POST'])
def import_deck_items(deck_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404

    body = request.get_json()
    items = body if isinstance(body, list) else body.get('items', [])

    existing = {di.item_order: di for di in DeckItem.query.filter_by(deck_id=deck_id).all()}
    max_order = max(existing) if existing else 0
    incoming_orders = set()
    next_free = max_order + 1

    # Cards present before this import that are NOT in the incoming set.
    # They will be removed as items, but their study progress is kept
    # (Progress/LearningEvent rows stay so statistics are not lost).
    before_count = len(existing)
    removed_orders = set(existing.keys())

    for i, item in enumerate(items):
        order = item.get('item_order')
        if order is None or order == 0:
            order = next_free
            next_free += 1
        else:
            order = int(order)
        incoming_orders.add(order)
        removed_orders.discard(order)

        if order in existing:
            di = existing[order]
            di.data = item.get('data', item)
            di.debug = item.get('debug', False)
        else:
            di = DeckItem(
                deck_id=deck_id,
                item_order=order,
                data=item.get('data', item),
                debug=item.get('debug', False),
            )
            db.session.add(di)

    for order in removed_orders:
        di = existing[order]
        # Remove the card itself, but keep Progress/LearningEvent history.
        db.session.delete(di)

    db.session.commit()
    return jsonify({
        'success': True,
        'count': len(items),
        'removed': len(removed_orders),
        'had_existing': before_count > 0,
    })


@app.route('/v1/decks/<int:deck_id>/import-excel', methods=['POST'])
def import_deck_items_excel(deck_id):
    d = db.session.get(Deck, deck_id)
    if not d:
        return jsonify({'error': 'Deck not found'}), 404

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No filename'}), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({'error': 'openpyxl not installed. Run: pip install openpyxl'}), 500

    try:
        wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'error': 'Excel file is empty'}), 400

        headers = [str(h).strip() if h is not None else f'col_{i}' for i, h in enumerate(rows[0])]

        # Replace all cards, but keep Progress/LearningEvent/StudyRound history
        # so study statistics survive the re-import.
        before_count = DeckItem.query.filter_by(deck_id=deck_id).count()
        DeckItem.query.filter_by(deck_id=deck_id).delete()

        count = 0
        for i, row in enumerate(rows[1:], start=1):
            if all(v is None for v in row):
                continue
            item_data = {}
            for j, val in enumerate(row):
                if j < len(headers) and val is not None:
                    item_data[headers[j]] = val if not isinstance(val, str) else val.strip()
            if item_data:
                di = DeckItem(
                    deck_id=deck_id,
                    item_order=i,
                    data=item_data,
                    debug=False,
                )
                db.session.add(di)
                count += 1

        db.session.commit()
        return jsonify({
            'success': True,
            'count': count,
            'removed': before_count,
            'had_existing': before_count > 0,
        })

    except Exception as e:
        return jsonify({'error': f'Excel parse error: {str(e)}'}), 400


# ==================== Learning ====================

def _init_progress(user_id, deck_id):
    """Ensure every deck item has a Progress record for this user (idempotent)."""
    items = DeckItem.query.filter_by(deck_id=deck_id).order_by(DeckItem.item_order).all()
    existing_ids = {
        p.deck_item_id
        for p in Progress.query.filter_by(user_id=user_id, deck_id=deck_id).all()
    }
    added = 0
    for item in items:
        if item.id in existing_ids:
            continue
        db.session.add(Progress(
            user_id=user_id, deck_id=deck_id, deck_item_id=item.id,
            is_unknown=0, current_order=item.item_order,
        ))
        added += 1
    if added:
        db.session.commit()


@app.route('/v1/learn/info')
def learn_info():
    user_id = request.args.get('user_id', type=int)
    deck_id = request.args.get('deck_id', type=int)

    total = DeckItem.query.filter_by(deck_id=deck_id).count()
    if user_id and deck_id:
        _init_progress(user_id, deck_id)
        unknown = Progress.query.filter_by(user_id=user_id, deck_id=deck_id, is_unknown=1)\
            .join(DeckItem, Progress.deck_item_id == DeckItem.id).count()
        known = total - unknown
    else:
        unknown = 0
        known = 0

    return jsonify({
        'total_words': total,
        'unknown_count': unknown,
        'known_count': known,
    })


@app.route('/v1/learn/page')
def learn_page():
    user_id = request.args.get('user_id', type=int)
    deck_id = request.args.get('deck_id', type=int)
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', Config.PAGE_SIZE, type=int)

    if not user_id or not deck_id:
        return jsonify({'error': 'user_id and deck_id required'}), 400

    _init_progress(user_id, deck_id)

    offset = (page - 1) * page_size
    # Join against existing deck items so orphaned Progress records
    # (from removed cards) are excluded from pagination.
    progress_query = Progress.query.filter_by(user_id=user_id, deck_id=deck_id)\
        .join(DeckItem, Progress.deck_item_id == DeckItem.id)\
        .order_by(Progress.current_order)
    total = progress_query.count()
    progress_list = progress_query.offset(offset).limit(page_size).all()

    cards_data = []
    for p in progress_list:
        item = db.session.get(DeckItem, p.deck_item_id)
        if item:
            d = item.to_dict()
            d['is_unknown'] = p.is_unknown
            d['is_favorite'] = p.is_favorite
            d['current_order'] = p.current_order
            cards_data.append(d)

    total_pages = math.ceil(total / page_size) if total else 0

    return jsonify({
        'cards': cards_data,
        'page': page,
        'page_size': page_size,
        'total': total,
        'total_pages': total_pages,
        'has_next': page < total_pages,
        'has_prev': page > 1,
    })


@app.route('/v1/learn/page_status')
def learn_page_status():
    user_id = request.args.get('user_id', type=int)
    deck_id = request.args.get('deck_id', type=int)
    page_size = Config.PAGE_SIZE

    unknown_count = Progress.query.filter_by(
        user_id=user_id, deck_id=deck_id, is_unknown=1
    ).join(DeckItem, Progress.deck_item_id == DeckItem.id).count()
    marked_pages = math.ceil(unknown_count / page_size) if unknown_count > 0 else 0

    return jsonify({
        'success': True,
        'marked_pages_count': marked_pages,
    })


@app.route('/v1/learn/mark', methods=['POST'])
def mark_item():
    data = request.get_json()
    deck_item_id = data.get('deck_item_id')
    user_id = data.get('user_id')
    deck_id = data.get('deck_id')
    is_unknown = data.get('is_unknown')

    if not all([deck_item_id, user_id, deck_id]):
        return jsonify({'error': 'deck_item_id, user_id, deck_id required'}), 400

    progress = Progress.query.filter_by(
        user_id=user_id, deck_id=deck_id, deck_item_id=deck_item_id
    ).first()
    if not progress:
        return jsonify({'error': 'Progress not found'}), 404

    if is_unknown is None:
        progress.is_unknown = 1 if progress.is_unknown == 0 else 0
    else:
        progress.is_unknown = 1 if is_unknown else 0

    db.session.commit()
    return jsonify({
        'success': True,
        'deck_item_id': deck_item_id,
        'is_unknown': progress.is_unknown,
    })


@app.route('/v1/reorder', methods=['POST'])
def reorder_cards():
    data = request.get_json() or {}
    user_id = data.get('user_id')
    deck_id = data.get('deck_id')

    if not user_id or not deck_id:
        return jsonify({'error': 'user_id and deck_id required'}), 400

    records = db.session.query(
        Progress.id, Progress.is_unknown, DeckItem.item_order
    ).join(DeckItem, Progress.deck_item_id == DeckItem.id).filter(
        Progress.user_id == user_id,
        Progress.deck_id == deck_id,
    ).order_by(
        Progress.is_unknown.desc(),
        DeckItem.item_order.asc(),
    ).all()

    for new_order, record in enumerate(records, start=1):
        p = db.session.get(Progress, record.id)
        p.current_order = new_order

    unknown_count = Progress.query.filter_by(
        user_id=user_id, deck_id=deck_id, is_unknown=1
    ).join(DeckItem, Progress.deck_item_id == DeckItem.id).count()

    max_round = db.session.query(db.func.max(StudyRound.round_number)).filter(
        StudyRound.user_id == user_id,
        StudyRound.deck_id == deck_id,
    ).scalar() or 0

    db.session.add(StudyRound(
        user_id=user_id, deck_id=deck_id,
        round_number=max_round + 1,
        end_time=datetime.utcnow(),
        marked_count=unknown_count,
    ))
    db.session.commit()

    return jsonify({
        'success': True,
        'total_cards': len(records),
    })


@app.route('/v1/rounds')
def get_rounds():
    user_id = request.args.get('user_id', type=int)
    deck_id = request.args.get('deck_id', type=int)
    q = StudyRound.query
    if user_id:
        q = q.filter_by(user_id=user_id)
    if deck_id:
        q = q.filter_by(deck_id=deck_id)
    rounds = q.order_by(StudyRound.round_number).all()
    return jsonify({'success': True, 'data': [r.to_dict() for r in rounds]})


@app.route('/v1/stats')
def get_stats():
    user_id = request.args.get('user_id', type=int)
    deck_id = request.args.get('deck_id', type=int)

    results = db.session.query(
        DeckItem.item_order, Progress.is_unknown
    ).join(Progress, DeckItem.id == Progress.deck_item_id).filter(
        Progress.user_id == user_id,
        Progress.deck_id == deck_id,
    ).all()

    data = [{'item_order': r.item_order, 'is_unknown': r.is_unknown} for r in results]
    return jsonify({'success': True, 'data': data})


@app.route('/v1/decks/<int:deck_id>/mastery')
def get_deck_mastery(deck_id):
    """Per-deck mastery: per-card state (0=mastered incl unstudied, 1=unknown) + study rounds."""
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({'error': 'user_id required'}), 400
    deck = db.session.get(Deck, deck_id)
    if not deck:
        return jsonify({'error': 'deck not found'}), 404

    items = db.session.query(DeckItem.item_order, DeckItem.id).filter(
        DeckItem.deck_id == deck_id
    ).order_by(DeckItem.item_order).all()

    progress_map = dict(db.session.query(
        Progress.deck_item_id, Progress.is_unknown
    ).filter(
        Progress.user_id == user_id,
        Progress.deck_id == deck_id,
    ).all())

    states = [progress_map.get(iid, 0) for _, iid in items]

    rounds = [
        {'round_number': r.round_number, 'end_time': r.end_time.isoformat() if r.end_time else None,
         'marked_count': r.marked_count}
        for r in db.session.query(StudyRound).filter(
            StudyRound.user_id == user_id,
            StudyRound.deck_id == deck_id,
        ).order_by(StudyRound.round_number).all()
    ]

    return jsonify({
        'success': True,
        'data': {
            'deck_name': deck.name,
            'item_count': len(items),
            'states': states,
            'rounds': rounds,
        }
    })


# ==================== Achievements ====================


@app.route('/v1/achievements')
def get_achievements():
    from datetime import date, timedelta
    user_id = request.args.get('user_id', type=int)
    if not user_id:
        return jsonify({'error': 'user_id required'}), 400

    deck_counts = db.session.query(
        StudyRound.deck_id, db.func.count(StudyRound.id)
    ).filter(StudyRound.user_id == user_id).group_by(StudyRound.deck_id).all()
    deck_names = {
        d.id: d.name for d in db.session.query(Deck).filter(Deck.user_id == user_id).all()
    }
    decks = [
        {'deck_id': did, 'name': deck_names.get(did, '卡组'), 'rounds': n}
        for did, n in deck_counts
    ]
    decks.sort(key=lambda d: d['rounds'], reverse=True)

    # 连续学习天数：按每天有学习事件（LearningEvent）判定
    study_dates = {r[0] for r in db.session.query(
        db.func.date(LearningEvent.created_at)
    ).filter(LearningEvent.user_id == user_id).distinct().all()}
    streak = 0
    anchor = date.today()
    if anchor.isoformat() not in study_dates:
        anchor = anchor - timedelta(days=1)
    while anchor.isoformat() in study_dates:
        streak += 1
        anchor = anchor - timedelta(days=1)

    deck_count = db.session.query(db.func.count(Deck.id)).filter(
        Deck.user_id == user_id
    ).scalar() or 0

    mastered_cards = db.session.query(db.func.count(Progress.id)).filter(
        Progress.user_id == user_id,
        Progress.is_unknown == 0,
    ).join(DeckItem, Progress.deck_item_id == DeckItem.id).scalar() or 0

    audio_play = db.session.query(db.func.count(LearningEvent.id)).filter(
        LearningEvent.user_id == user_id,
        LearningEvent.action == 'audio_play',
    ).scalar() or 0

    word_mark = db.session.query(db.func.count(LearningEvent.id)).filter(
        LearningEvent.user_id == user_id,
        LearningEvent.action == 'word_mark',
    ).scalar() or 0

    return jsonify({
        'success': True,
        'data': {
            'decks': decks,
            'streak_days': streak,
            'deck_count': deck_count,
            'mastered_cards': mastered_cards,
            'audio_play': audio_play,
            'word_mark': word_mark,
        }
    })


# ==================== Observability ====================


@app.route('/v1/observability/event', methods=['POST'])
def record_event():
    data = request.get_json()
    user_id = data.get('user_id')
    deck_id = data.get('deck_id')
    deck_item_id = data.get('deck_item_id')
    action = data.get('action')
    template_id = data.get('template_id')

    if not all([user_id, deck_id, deck_item_id, action]):
        return jsonify({'error': 'user_id, deck_id, deck_item_id, action required'}), 400
    if not isinstance(action, str) or not action.strip():
        return jsonify({'error': 'action must be a non-empty string'}), 400

    event = LearningEvent(
        user_id=user_id, deck_id=deck_id, deck_item_id=deck_item_id,
        template_id=template_id, action=action.strip(),
    )
    db.session.add(event)
    db.session.commit()
    return jsonify({'success': True, 'event_id': event.id})


@app.route('/v1/observability/events', methods=['POST'])
def record_events_batch():
    data = request.get_json()
    events = data.get('events') or []
    count = 0
    for e in events:
        user_id = e.get('user_id')
        deck_id = e.get('deck_id')
        deck_item_id = e.get('deck_item_id')
        action = e.get('action')
        template_id = e.get('template_id')
        if all([user_id, deck_id, deck_item_id, action]) and isinstance(action, str) and action.strip():
            db.session.add(LearningEvent(
                user_id=user_id, deck_id=deck_id, deck_item_id=deck_item_id,
                template_id=template_id, action=action.strip(),
            ))
            count += 1
    db.session.commit()
    return jsonify({'success': True, 'count': count})


@app.route('/v1/observability/actions')
def get_observability_actions():
    """Return action types from template definition, falling back to distinct DB events."""
    deck_id = request.args.get('deck_id', type=int)
    template_id = request.args.get('template_id', type=int)

    if deck_id:
        d = db.session.get(Deck, deck_id)
        if d:
            tid = d.active_template_id
            if tid:
                t = db.session.get(Template, tid)
                if t and t.tracked_actions:
                    try:
                        actions = json.loads(t.tracked_actions)
                        return jsonify({'success': True, 'actions': actions})
                    except json.JSONDecodeError:
                        pass
    if template_id:
        t = db.session.get(Template, template_id)
        if t and t.tracked_actions:
            try:
                actions = json.loads(t.tracked_actions)
                return jsonify({'success': True, 'actions': actions})
            except json.JSONDecodeError:
                pass

    q = db.session.query(LearningEvent.action).distinct()
    if deck_id:
        q = q.filter(LearningEvent.deck_id == deck_id)
    elif template_id:
        q = q.filter(LearningEvent.template_id == template_id)
    actions = [r[0] for r in q.order_by(LearningEvent.action).all()]
    return jsonify({'success': True, 'actions': actions})


@app.route('/v1/observability/data')
def get_observability_data():
    from datetime import datetime, timedelta
    user_id = request.args.get('user_id', type=int)
    deck_id = request.args.get('deck_id', type=int)
    template_id = request.args.get('template_id', type=int)
    view = request.args.get('view', 'daily')
    date_str = request.args.get('date')

    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format'}), 400
    else:
        target_date = datetime.now().date()

    if view == 'daily':
        start_dt = datetime.combine(target_date, datetime.min.time())
        end_dt = start_dt + timedelta(days=1)
    elif view == 'weekly':
        start_of_week = target_date - timedelta(days=target_date.weekday())
        start_dt = datetime.combine(start_of_week, datetime.min.time())
        end_dt = start_dt + timedelta(weeks=1)
    elif view == 'monthly':
        start_of_month = target_date.replace(day=1)
        start_dt = datetime.combine(start_of_month, datetime.min.time())
        if target_date.month == 12:
            end_dt = datetime(target_date.year + 1, 1, 1)
        else:
            end_dt = datetime(target_date.year, target_date.month + 1, 1)
    elif view == 'heatmap':
        start_dt = datetime.combine(target_date - timedelta(days=364), datetime.min.time())
        end_dt = datetime.combine(target_date + timedelta(days=1), datetime.min.time())
    else:
        return jsonify({'error': 'Invalid view'}), 400

    q = LearningEvent.query.filter(
        LearningEvent.created_at >= start_dt,
        LearningEvent.created_at < end_dt,
    )
    if user_id:
        q = q.filter(LearningEvent.user_id == user_id)
    if deck_id:
        q = q.filter(LearningEvent.deck_id == deck_id)
    if template_id:
        q = q.filter(LearningEvent.template_id == template_id)
    events = q.order_by(LearningEvent.created_at).all()

    from collections import defaultdict

    if view == 'heatmap':
        heatmap = defaultdict(lambda: defaultdict(int))
        for ev in events:
            date_key = ev.created_at.date().isoformat()
            heatmap[date_key][ev.action] += 1
        result = [{'date': d, 'actions': dict(a)}
                   for d, a in sorted(heatmap.items())]
    elif view == 'daily':
        bucket_stats = defaultdict(lambda: defaultdict(int))
        for ev in events:
            minute = ev.created_at.minute
            bucket_minute = (minute // 15) * 15
            key = f"{ev.created_at.hour:02d}:{bucket_minute:02d}"
            bucket_stats[key][ev.action] += 1
        result = [{'time': k, 'actions': dict(v)}
                   for k, v in sorted(bucket_stats.items())]
    else:
        daily = defaultdict(lambda: defaultdict(int))
        for ev in events:
            date_key = ev.created_at.date().isoformat()
            daily[date_key][ev.action] += 1
        result = [{'date': d, 'actions': dict(a)}
                   for d, a in sorted(daily.items())]

    return jsonify({
        'success': True,
        'view': view,
        'template_id': template_id,
        'date_range': {
            'start': start_dt.date().isoformat(),
            'end': (end_dt - timedelta(days=1)).date().isoformat(),
        },
        'data': result,
    })


# ==================== Favorite ====================

@app.route('/v1/learn/favorite', methods=['POST'])
def toggle_favorite():
    data = request.get_json()
    deck_item_id = data.get('deck_item_id')
    user_id = data.get('user_id')
    deck_id = data.get('deck_id')

    if not all([deck_item_id, user_id, deck_id]):
        return jsonify({'error': 'deck_item_id, user_id, deck_id required'}), 400

    progress = Progress.query.filter_by(
        user_id=user_id, deck_id=deck_id, deck_item_id=deck_item_id
    ).first()
    if not progress:
        return jsonify({'error': 'Progress not found'}), 404

    progress.is_favorite = 0 if progress.is_favorite else 1
    db.session.commit()
    return jsonify({
        'success': True,
        'deck_item_id': deck_item_id,
        'is_favorite': progress.is_favorite,
    })


# ==================== Page Routes ====================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/v1/templates/<int:template_id>/preview')
def preview_template_cards(template_id):
    """Return sample card data for template preview."""
    t = db.session.get(Template, template_id)
    if not t:
        return jsonify({'error': 'Template not found'}), 404

    d = Deck.query.filter(Deck.active_template_id == template_id).first()
    sample_item = None
    if d:
        item = DeckItem.query.filter_by(deck_id=d.id).order_by(DeckItem.item_order).first()
        if item:
            sample_item = item.to_dict()

    return jsonify({
        'success': True,
        'template': t.to_dict_full(),
        'sample_card': sample_item,
    })


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001)
